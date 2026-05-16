import io
from typing import List, Optional
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, text
from sqlalchemy.orm import Session

from app import models, schemas, auth
from app.database import get_db

router = APIRouter()


@router.get("/summary", response_model=List[schemas.MonthSummary])
def monthly_summary(
    months: int = Query(6, le=24),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    rows = (
        db.query(
            models.Transaction.periodo_referencia,
            models.Transaction.type,
            func.sum(models.Transaction.amount).label("total"),
        )
        .filter(models.Transaction.user_id == current_user.id)
        .group_by(models.Transaction.periodo_referencia, models.Transaction.type)
        .order_by(models.Transaction.periodo_referencia.desc())
        .all()
    )

    # Aggregate by period
    periods: dict = {}
    for row in rows:
        p = row.periodo_referencia
        if p not in periods:
            periods[p] = {"periodo_referencia": p, "total_receitas": 0, "total_despesas": 0}
        if row.type == "R":
            periods[p]["total_receitas"] = round(row.total, 2)
        else:
            periods[p]["total_despesas"] = round(row.total, 2)

    result = []
    for p, d in sorted(periods.items(), reverse=True)[:months]:
        d["saldo"] = round(d["total_receitas"] - d["total_despesas"], 2)
        result.append(schemas.MonthSummary(**d))
    return result


@router.get("/comparative")
def comparative(
    current_periodo: str = Query(..., description="YYYY-MM"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    """Compare current month vs previous month, broken down by IFRS category."""

    def _get_summary(periodo: str):
        rows = (
            db.query(
                models.Transaction.type,
                func.sum(models.Transaction.amount).label("total"),
            )
            .filter(
                models.Transaction.user_id == current_user.id,
                models.Transaction.periodo_referencia == periodo,
            )
            .group_by(models.Transaction.type)
            .all()
        )
        rec = sum(r.total for r in rows if r.type == "R")
        des = sum(r.total for r in rows if r.type == "D")
        return {"periodo_referencia": periodo, "total_receitas": round(rec, 2),
                "total_despesas": round(des, 2), "saldo": round(rec - des, 2)}

    def _prev(periodo: str) -> str:
        y, m = int(periodo[:4]), int(periodo[5:])
        m -= 1
        if m == 0:
            m, y = 12, y - 1
        return f"{y:04d}-{m:02d}"

    prev_periodo = _prev(current_periodo)
    current_data = _get_summary(current_periodo)
    previous_data = _get_summary(prev_periodo)

    var = 0.0
    if previous_data["total_despesas"] > 0:
        var = ((current_data["total_despesas"] - previous_data["total_despesas"]) /
               previous_data["total_despesas"]) * 100

    # By category
    cat_rows = (
        db.query(
            models.Category.name,
            models.Category.ifrs_group,
            models.Transaction.periodo_referencia,
            func.sum(models.Transaction.amount).label("total"),
        )
        .join(models.Transaction, models.Transaction.category_id == models.Category.id)
        .filter(
            models.Transaction.user_id == current_user.id,
            models.Transaction.type == "D",
            models.Transaction.periodo_referencia.in_([current_periodo, prev_periodo]),
        )
        .group_by(models.Category.name, models.Category.ifrs_group, models.Transaction.periodo_referencia)
        .all()
    )

    cat_map: dict = {}
    for row in cat_rows:
        if row.name not in cat_map:
            cat_map[row.name] = {"name": row.name, "ifrs": row.ifrs_group, "current": 0, "previous": 0}
        if row.periodo_referencia == current_periodo:
            cat_map[row.name]["current"] = round(row.total, 2)
        else:
            cat_map[row.name]["previous"] = round(row.total, 2)

    return {
        "current": current_data,
        "previous": previous_data,
        "variation_pct": round(var, 2),
        "by_category": list(cat_map.values()),
    }


@router.get("/by-card")
def by_card(
    periodo: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    q = (
        db.query(
            models.Card.name,
            models.Card.bank,
            func.sum(models.Transaction.amount).label("total"),
            func.count(models.Transaction.id).label("count"),
        )
        .join(models.Transaction, models.Transaction.card_id == models.Card.id)
        .filter(
            models.Transaction.user_id == current_user.id,
            models.Transaction.type == "D",
        )
    )
    if periodo:
        q = q.filter(models.Transaction.periodo_referencia == periodo)
    q = q.group_by(models.Card.name, models.Card.bank).all()
    return [{"card": r.name, "bank": r.bank, "total": round(r.total, 2), "count": r.count} for r in q]


@router.get("/installments-projection")
def installments_projection(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    rows = (
        db.query(
            models.Transaction.installment_group,
            models.Transaction.description,
            models.Transaction.amount,
            func.count(models.Transaction.id).label("remaining"),
            func.min(models.Transaction.date).label("next_date"),
        )
        .filter(
            models.Transaction.user_id == current_user.id,
            models.Transaction.installment_group.isnot(None),
            models.Transaction.type == "D",
        )
        .group_by(
            models.Transaction.installment_group,
            models.Transaction.description,
            models.Transaction.amount,
        )
        .all()
    )
    return [
        {
            "description": r.description,
            "monthly_amount": round(r.amount, 2),
            "remaining": r.remaining,
            "next_date": str(r.next_date),
            "total_committed": round(r.amount * r.remaining, 2),
        }
        for r in rows
    ]


def _compute_dre_data(periodo: str, user_id: int, db: Session) -> dict:
    IFRS_ORDER = [
        "Custos Operacionais",
        "Despesas Administrativas",
        "Despesas Fixas",
        "Despesas Financeiras",
        "Investimentos (Ativos)",
    ]

    receita_bruta = round(float(
        db.query(func.sum(models.Transaction.amount))
        .filter(
            models.Transaction.user_id == user_id,
            models.Transaction.periodo_referencia == periodo,
            models.Transaction.type == "R",
        ).scalar() or 0
    ), 2)

    cat_rows = (
        db.query(
            models.Category.name.label("cat_name"),
            models.Category.ifrs_group,
            models.Category.icon,
            models.Category.color,
            func.sum(models.Transaction.amount).label("total"),
        )
        .join(models.Transaction, models.Transaction.category_id == models.Category.id)
        .filter(
            models.Transaction.user_id == user_id,
            models.Transaction.periodo_referencia == periodo,
            models.Transaction.type == "D",
        )
        .group_by(
            models.Category.name,
            models.Category.ifrs_group,
            models.Category.icon,
            models.Category.color,
        )
        .all()
    )

    uncat = round(float(
        db.query(func.sum(models.Transaction.amount))
        .filter(
            models.Transaction.user_id == user_id,
            models.Transaction.periodo_referencia == periodo,
            models.Transaction.type == "D",
            models.Transaction.category_id.is_(None),
        ).scalar() or 0
    ), 2)

    group_map = {g: {"name": g, "total": 0.0, "categories": []} for g in IFRS_ORDER}

    for row in cat_rows:
        g = row.ifrs_group
        if g not in group_map:
            group_map[g] = {"name": g, "total": 0.0, "categories": []}
        group_map[g]["categories"].append({
            "name": row.cat_name, "icon": row.icon,
            "color": row.color, "total": round(float(row.total), 2),
        })
        group_map[g]["total"] = round(group_map[g]["total"] + float(row.total), 2)

    if uncat > 0:
        outros = "Outros"
        if outros not in group_map:
            group_map[outros] = {"name": outros, "total": 0.0, "categories": []}
        group_map[outros]["categories"].append(
            {"name": "Sem categoria", "icon": "❓", "color": "#888888", "total": uncat}
        )
        group_map[outros]["total"] = round(group_map[outros]["total"] + uncat, 2)

    grupos = [group_map[g] for g in IFRS_ORDER]
    for g, data in group_map.items():
        if g not in IFRS_ORDER:
            grupos.append(data)

    total_despesas = round(sum(g["total"] for g in grupos), 2)
    return {
        "periodo": periodo,
        "receita_bruta": receita_bruta,
        "grupos": grupos,
        "total_despesas": total_despesas,
        "resultado": round(receita_bruta - total_despesas, 2),
    }


@router.get("/dre")
def dre(
    periodo: str = Query(..., description="YYYY-MM"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    return _compute_dre_data(periodo, current_user.id, db)


@router.get("/export/dre")
def export_dre(
    periodo: str = Query(..., description="YYYY-MM"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import datetime

    data = _compute_dre_data(periodo, current_user.id, db)
    receita = data["receita_bruta"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DRE"
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 12

    cur = [1]

    def pct(amount):
        return round(amount / receita * 100, 1) if receita > 0 else 0.0

    STYLES = {
        "receita":    ("D1FAE5", "065F46", True,  12),
        "group":      ("FFF5F5", "9B2335", True,  11),
        "category":   ("F9F9F9", "555555", False, 10),
        "subtotal":   ("EFF6FF", "1D4ED8", True,  11),
        "result_pos": ("D1FAE5", "065F46", True,  13),
        "result_neg": ("FEF2F2", "991B1B", True,  13),
    }

    def wr(label, amount, pct_val=None, indent=0, style="normal"):
        r = cur[0]
        ws.row_dimensions[r].height = 24 if style in ("receita", "subtotal", "result_pos", "result_neg") else 17
        lc = ws.cell(row=r, column=2, value=("    " * indent) + label)
        vc = ws.cell(row=r, column=3, value=round(amount, 2))
        vc.number_format = '"R$" #,##0.00'
        vc.alignment = Alignment(horizontal="right")
        if pct_val is not None:
            pc = ws.cell(row=r, column=4, value=pct_val / 100)
            pc.number_format = "0.0%"
            pc.alignment = Alignment(horizontal="right")
        if style in STYLES:
            bg, color, bold, sz = STYLES[style]
            for col in range(2, 5):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=bg)
            lc.font = Font(bold=bold, size=sz, color=color)
            vc.font = Font(bold=bold, size=sz, color=color)
            if pct_val is not None:
                ws.cell(row=r, column=4).font = Font(bold=bold, size=sz, color=color)
        cur[0] += 1

    # Title rows
    r = cur[0]
    ws.merge_cells(f"B{r}:D{r}")
    ws.cell(row=r, column=2).value = "DRE — Demonstração do Resultado do Exercício"
    ws.cell(row=r, column=2).font = Font(bold=True, size=14, color="9B2335")
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
    ws.row_dimensions[r].height = 30
    cur[0] += 1
    r = cur[0]
    ws.merge_cells(f"B{r}:D{r}")
    ws.cell(row=r, column=2).value = (
        f"Período: {periodo}  |  {current_user.name}  |  {datetime.date.today().strftime('%d/%m/%Y')}"
    )
    ws.cell(row=r, column=2).font = Font(size=10, color="888888", italic=True)
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
    cur[0] += 2

    SECTIONS = [
        {"subtotal": "LUCRO BRUTO",           "groups": ["Custos Operacionais"]},
        {"subtotal": "RESULTADO OPERACIONAL",  "groups": ["Despesas Administrativas", "Despesas Fixas", "Despesas Financeiras"]},
        {"subtotal": "RESULTADO DO EXERCÍCIO", "groups": ["Investimentos (Ativos)"], "final": True},
    ]

    grupo_map = {g["name"]: g for g in data["grupos"]}
    wr("(+) RECEITA BRUTA DE VENDAS", receita, pct(receita), style="receita")
    cur[0] += 1
    running = receita

    for section in SECTIONS:
        for g_name in section["groups"]:
            g = grupo_map.get(g_name, {"name": g_name, "total": 0.0, "categories": []})
            wr(f"(-) {g['name']}", g["total"], pct(g["total"]), style="group")
            for cat in sorted(g.get("categories", []), key=lambda x: -x["total"]):
                wr(f"{cat['icon']} {cat['name']}", cat["total"], pct(cat["total"]), indent=1, style="category")
            running -= g["total"]
        style = ("result_pos" if running >= 0 else "result_neg") if section.get("final") else "subtotal"
        wr(f"(=) {section['subtotal']}", running, pct(running), style=style)
        if not section.get("final"):
            cur[0] += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=dre_{periodo}.xlsx"},
    )


@router.get("/export/reconciliation")
def export_reconciliation(
    periodo: str = Query(..., description="YYYY-MM"),
    card_id: Optional[int] = Query(None),
    category_ids: Optional[str] = Query(None, description="comma-separated IDs; empty=all"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from collections import OrderedDict
    import datetime

    q = db.query(models.Transaction).filter(
        models.Transaction.user_id == current_user.id,
        models.Transaction.periodo_referencia == periodo,
        models.Transaction.type == "D",
    )
    if card_id:
        q = q.filter(models.Transaction.card_id == card_id)
    if category_ids and category_ids.strip():
        ids = [int(x) for x in category_ids.split(",") if x.strip().isdigit()]
        if ids:
            q = q.filter(models.Transaction.category_id.in_(ids))

    txs = q.order_by(models.Transaction.category_id.nullslast(), models.Transaction.date).all()

    RED, GOLD = "9B2335", "C97B3C"
    DARK, MID = "1C110A", "231509"

    wb = openpyxl.Workbook()

    # ── Sheet 1: detailed transactions ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Conciliação"

    for col, w in zip("ABCDEFGH", [12, 36, 22, 18, 10, 25, 24, 16]):
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[3].height = 20

    # Title
    ws["A1"].value = f"Relatório de Conciliação — {periodo}"
    ws["A1"].font = Font(bold=True, size=14, color=RED)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:H1")

    ws["A2"].value = (
        f"Gerado em: {datetime.date.today().strftime('%d/%m/%Y')}  |  Usuário: {current_user.name}"
    )
    ws["A2"].font = Font(size=9, color="888888", italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:H2")

    # Column headers
    for col_idx, h in enumerate(
        ["Data", "Descrição", "Fornecedor", "Cartão", "Parcela", "Categoria", "Grupo IFRS", "Valor (R$)"], 1
    ):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.fill = PatternFill("solid", fgColor=RED)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Group transactions by category
    groups: OrderedDict = OrderedDict()
    for tx in txs:
        cat_name = tx.category.name if tx.category else "Sem categoria"
        ifrs_grp = tx.category.ifrs_group if tx.category else "—"
        key = (cat_name, ifrs_grp)
        if key not in groups:
            groups[key] = []
        groups[key].append(tx)

    row = 4
    grand_total = 0.0

    for (cat_name, ifrs_grp), group_txs in groups.items():
        cat_subtotal = sum(t.amount for t in group_txs)
        grand_total += cat_subtotal

        # Category header row
        ws.row_dimensions[row].height = 18
        ws.merge_cells(f"A{row}:G{row}")
        ws.cell(row=row, column=1).value = f"  {cat_name}  ·  {ifrs_grp}"
        ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF", size=10)
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=GOLD)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")

        sub_cell = ws.cell(row=row, column=8, value=round(cat_subtotal, 2))
        sub_cell.number_format = 'R$ #,##0.00'
        sub_cell.font = Font(bold=True, color="FFFFFF", size=10)
        sub_cell.fill = PatternFill("solid", fgColor=GOLD)
        sub_cell.alignment = Alignment(horizontal="right", vertical="center")
        row += 1

        for i, tx in enumerate(group_txs):
            ws.row_dimensions[row].height = 15
            bg = "FFF8F0" if i % 2 == 0 else "F5E6D3"

            ws.cell(row=row, column=1, value=tx.date.strftime("%d/%m/%Y"))
            ws.cell(row=row, column=2, value=tx.description)
            ws.cell(row=row, column=3, value=tx.supplier or "—")
            ws.cell(row=row, column=4, value=tx.card.name if tx.card else "—")
            inst = (
                f"{tx.installment_current}/{tx.installment_total}"
                if tx.installment_current else "À vista"
            )
            ws.cell(row=row, column=5, value=inst)
            ws.cell(row=row, column=6, value=cat_name)
            ws.cell(row=row, column=7, value=ifrs_grp)
            val_cell = ws.cell(row=row, column=8, value=round(tx.amount, 2))
            val_cell.number_format = 'R$ #,##0.00'

            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=bg)
                ws.cell(row=row, column=col).font = Font(size=10)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=8).alignment = Alignment(horizontal="right")
            row += 1

    # Grand total
    row += 1
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"A{row}:G{row}")
    ws.cell(row=row, column=1, value="TOTAL GERAL")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    total_cell = ws.cell(row=row, column=8, value=round(grand_total, 2))
    total_cell.number_format = 'R$ #,##0.00'
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    for col in range(1, 9):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=RED)
        ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF", size=12)

    # ── Sheet 2: summary by category ───────────────────────────────────────────
    ws2 = wb.create_sheet("Resumo")
    for col, w in zip("ABCD", [28, 28, 8, 16]):
        ws2.column_dimensions[col].width = w

    ws2["A1"].value = "Resumo por Categoria"
    ws2["A1"].font = Font(bold=True, size=13, color=RED)
    ws2.merge_cells("A1:D1")
    ws2["A1"].alignment = Alignment(horizontal="center")

    for col_idx, h in enumerate(["Categoria", "Grupo IFRS", "Qtd", "Total (R$)"], 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.fill = PatternFill("solid", fgColor=RED)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center")

    r2 = 4
    for (cat_name, ifrs_grp), group_txs in groups.items():
        bg = "FFF8F0" if r2 % 2 == 0 else "F5E6D3"
        ws2.cell(row=r2, column=1, value=cat_name)
        ws2.cell(row=r2, column=2, value=ifrs_grp)
        ws2.cell(row=r2, column=3, value=len(group_txs))
        v = ws2.cell(row=r2, column=4, value=round(sum(t.amount for t in group_txs), 2))
        v.number_format = 'R$ #,##0.00'
        for col in range(1, 5):
            ws2.cell(row=r2, column=col).fill = PatternFill("solid", fgColor=bg)
            ws2.cell(row=r2, column=col).font = Font(size=10)
        ws2.cell(row=r2, column=3).alignment = Alignment(horizontal="center")
        ws2.cell(row=r2, column=4).alignment = Alignment(horizontal="right")
        r2 += 1

    r2 += 1
    ws2.merge_cells(f"A{r2}:C{r2}")
    ws2.cell(row=r2, column=1, value="TOTAL")
    t2 = ws2.cell(row=r2, column=4, value=round(grand_total, 2))
    t2.number_format = 'R$ #,##0.00'
    for col in range(1, 5):
        ws2.cell(row=r2, column=col).fill = PatternFill("solid", fgColor=RED)
        ws2.cell(row=r2, column=col).font = Font(bold=True, color="FFFFFF", size=11)
    ws2.cell(row=r2, column=1).alignment = Alignment(horizontal="center")
    ws2.cell(row=r2, column=4).alignment = Alignment(horizontal="right")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=conciliacao_{periodo}.xlsx"},
    )


@router.get("/export/excel")
def export_excel(
    periodo: str = Query(..., description="YYYY-MM"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def _get_summary_data():
        rows = db.query(
            models.Transaction.type,
            func.sum(models.Transaction.amount).label("total"),
        ).filter(
            models.Transaction.user_id == current_user.id,
            models.Transaction.periodo_referencia == periodo,
        ).group_by(models.Transaction.type).all()
        rec = sum(r.total for r in rows if r.type == "R")
        des = sum(r.total for r in rows if r.type == "D")
        return rec, des

    def _get_cat_data():
        return db.query(
            models.Category.name,
            models.Category.ifrs_group,
            func.sum(models.Transaction.amount).label("total"),
        ).join(models.Transaction, models.Transaction.category_id == models.Category.id).filter(
            models.Transaction.user_id == current_user.id,
            models.Transaction.type == "D",
            models.Transaction.periodo_referencia == periodo,
        ).group_by(models.Category.name, models.Category.ifrs_group).order_by(func.sum(models.Transaction.amount).desc()).all()

    def _get_transactions():
        return db.query(models.Transaction).filter(
            models.Transaction.user_id == current_user.id,
            models.Transaction.periodo_referencia == periodo,
        ).order_by(models.Transaction.date).all()

    rec, des = _get_summary_data()
    cat_rows = _get_cat_data()
    txs = _get_transactions()

    wb = openpyxl.Workbook()

    # ── Sheet 1: Resumo ──
    ws1 = wb.active
    ws1.title = "Resumo"
    header_fill = PatternFill("solid", fgColor="9B2335")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 20

    ws1["A1"] = f"Relatório Financeiro — {periodo}"
    ws1["A1"].font = Font(bold=True, size=14, color="9B2335")
    ws1.merge_cells("A1:B1")
    ws1["A1"].alignment = Alignment(horizontal="center")

    ws1["A3"] = "Indicador"
    ws1["B3"] = "Valor (R$)"
    for cell in [ws1["A3"], ws1["B3"]]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    data = [
        ("Total Receitas", rec),
        ("Total Despesas", des),
        ("Saldo", rec - des),
    ]
    for i, (label, value) in enumerate(data, start=4):
        ws1[f"A{i}"] = label
        ws1[f"B{i}"] = round(value, 2)
        ws1[f"B{i}"].number_format = 'R$ #,##0.00'

    ws1["A8"] = "Por Categoria"
    ws1["A8"].font = Font(bold=True, size=12, color="9B2335")
    ws1.merge_cells("A8:B8")
    ws1["A9"] = "Categoria"
    ws1["B9"] = "Total (R$)"
    ws1["C9"] = "Grupo IFRS"
    ws1.column_dimensions["C"].width = 30
    for cell in [ws1["A9"], ws1["B9"], ws1["C9"]]:
        cell.fill = header_fill
        cell.font = header_font

    for i, row in enumerate(cat_rows, start=10):
        ws1[f"A{i}"] = row.name
        ws1[f"B{i}"] = round(row.total, 2)
        ws1[f"B{i}"].number_format = 'R$ #,##0.00'
        ws1[f"C{i}"] = row.ifrs_group

    # ── Sheet 2: Lançamentos ──
    ws2 = wb.create_sheet("Lançamentos")
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 35
    ws2.column_dimensions["D"].width = 20
    ws2.column_dimensions["E"].width = 15
    ws2.column_dimensions["F"].width = 25

    headers = ["Data", "Tipo", "Descrição", "Fornecedor", "Valor (R$)", "Categoria"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for i, tx in enumerate(txs, start=2):
        ws2.cell(row=i, column=1, value=str(tx.date))
        ws2.cell(row=i, column=2, value="Receita" if tx.type == "R" else "Despesa")
        ws2.cell(row=i, column=3, value=tx.description)
        ws2.cell(row=i, column=4, value=tx.supplier or "")
        cell_val = ws2.cell(row=i, column=5, value=round(tx.amount, 2))
        cell_val.number_format = 'R$ #,##0.00'
        ws2.cell(row=i, column=6, value=tx.category.name if tx.category else "Sem categoria")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=relatorio_{periodo}.xlsx"},
    )


@router.get("/export/pdf")
def export_pdf(
    periodo: str = Query(..., description="YYYY-MM"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, HRFlowable
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    def _get_summary_data():
        rows = db.query(
            models.Transaction.type,
            func.sum(models.Transaction.amount).label("total"),
        ).filter(
            models.Transaction.user_id == current_user.id,
            models.Transaction.periodo_referencia == periodo,
        ).group_by(models.Transaction.type).all()
        rec = sum(r.total for r in rows if r.type == "R")
        des = sum(r.total for r in rows if r.type == "D")
        return rec, des

    def _get_cat_data():
        return db.query(
            models.Category.name,
            models.Category.ifrs_group,
            func.sum(models.Transaction.amount).label("total"),
        ).join(models.Transaction, models.Transaction.category_id == models.Category.id).filter(
            models.Transaction.user_id == current_user.id,
            models.Transaction.type == "D",
            models.Transaction.periodo_referencia == periodo,
        ).group_by(models.Category.name, models.Category.ifrs_group).order_by(func.sum(models.Transaction.amount).desc()).all()

    def _get_transactions():
        return db.query(models.Transaction).filter(
            models.Transaction.user_id == current_user.id,
            models.Transaction.periodo_referencia == periodo,
        ).order_by(models.Transaction.date).all()

    rec, des = _get_summary_data()
    cat_rows = _get_cat_data()
    txs = _get_transactions()

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm
    )

    ACCENT = colors.HexColor("#9B2335")
    GOLD = colors.HexColor("#C97B3C")
    DARK = colors.HexColor("#1C110A")
    LIGHT_GRAY = colors.HexColor("#F5F5F5")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=18, textColor=ACCENT, spaceAfter=6)
    sub_style   = ParagraphStyle("sub",   parent=styles["Normal"], fontSize=10, textColor=colors.gray, spaceAfter=12)
    h2_style    = ParagraphStyle("h2",    parent=styles["Heading2"], fontSize=13, textColor=ACCENT, spaceBefore=14, spaceAfter=8)
    normal      = styles["Normal"]

    story = []
    story.append(Paragraph("FinanceControl", title_style))
    story.append(Paragraph(f"Relatório Financeiro — {periodo} | {current_user.name}", sub_style))
    story.append(HRFlowable(width="100%", thickness=1, color=ACCENT, spaceAfter=14))

    # Summary table
    story.append(Paragraph("Resumo do Período", h2_style))
    summary_data = [
        ["Indicador", "Valor"],
        ["Total Receitas", f"R$ {rec:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")],
        ["Total Despesas", f"R$ {des:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")],
        ["Saldo", f"R$ {rec-des:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")],
    ]
    t = Table(summary_data, colWidths=[10*cm, 6*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), ACCENT),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t)

    if cat_rows:
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph("Despesas por Categoria", h2_style))
        cat_table_data = [["Categoria", "Grupo IFRS", "Total"]]
        for row in cat_rows:
            val = f"R$ {row.total:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            cat_table_data.append([row.name, row.ifrs_group, val])
        ct = Table(cat_table_data, colWidths=[6*cm, 7*cm, 4*cm])
        ct.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), ACCENT),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(ct)

    if txs:
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph("Lançamentos do Período", h2_style))
        tx_data = [["Data", "Tipo", "Descrição", "Valor"]]
        for tx in txs:
            val = f"R$ {tx.amount:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            tx_data.append([
                str(tx.date),
                "Receita" if tx.type == "R" else "Despesa",
                tx.description[:40] + ("..." if len(tx.description) > 40 else ""),
                val,
            ])
        txt = Table(tx_data, colWidths=[2.5*cm, 2.5*cm, 9*cm, 3*cm])
        txt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), ACCENT),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
            ("ALIGN", (3, 0), (3, -1), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(txt)

    doc.build(story)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=relatorio_{periodo}.pdf"},
    )
